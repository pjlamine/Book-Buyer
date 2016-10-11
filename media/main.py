import bookFinder
from openpyxl import Workbook
from openpyxl import load_workbook
#
#
wb2 = load_workbook('Summer 16.xlsx')
sheetNames = wb2.get_sheet_names()
ws2 = wb2.get_sheet_by_name(sheetNames[0])

for a in range (2,42):
    # Inventory [ws2.cell('A' + str(a)).value] = ws2.cell('C' + str(a)).value
    # print ws2.cell('A' + str(a)).value
    isbn = ws2.cell('A' + str(a)).value

    try:
        print isbn
        # book[a] = BookScouter.BookScouter(isbn)
        sellPrice = bookFinder.getPrice(isbn, 2)
        ws2['G' + str(a)] = sellPrice

    except Exception as e:
        print "Caught exception "
        print e
        # ws2['D' + str(a)] = "Error"
        # ws2['E' + str(a)] = "Error"

        continue

wb2.save("Summer 16.xlsx")

# isbn = 9780321757272
#
# scout = BookScouter.BookScouter(isbn)
# prices = scout.getPrice()
# #
# buyPrice = prices[0]
# cheggPrice = prices[1]
# print prices
