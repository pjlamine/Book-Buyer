import BookScouter
from openpyxl import Workbook
from openpyxl import load_workbook


wb2 = load_workbook('Summer 16.xlsx')
sheetNames = wb2.get_sheet_names()
ws2 = wb2.get_sheet_by_name(sheetNames[0])

Inventory = {}

for a in range (2,42):
    # Inventory [ws2.cell('A' + str(a)).value] = ws2.cell('C' + str(a)).value
    # print ws2.cell('A' + str(a)).value
    isbn = ws2.cell('A' + str(a)).value
    book = {}
    try:
        print isbn
        book[a] = BookScouter.BookScouter(isbn)
        prices = book[a].getPrice()                 #returns list of [buyPrice, cheggPrice, title]
        print prices
        buyPrice = prices[0]
        cheggPrice = prices[1]
        title = prices[2]
        ws2['D' + str(a)] = buyPrice
        ws2['E' + str(a)] = cheggPrice
        ws2['B' + str(a)] = title

    except Exception as e:
        print "Caught exception "
        print e
        ws2['D' + str(a)] = "Error"
        ws2['E' + str(a)] = "Error"

        continue
# for j in Inventory:
#     print j , Inventory[j]
# book =  BookScouter.BookScouter("9781107612495")
#
# scoutPrices=  book.getPrice()
#
# buyPrice = scoutPrices[0]
# cheggPrice = scoutPrices[1]
#
# print "Buying Price:" ,buyPrice , "\nCheggPrice" , cheggPrice




#
# print BookScouter.getPrice("9781107612495")
#
# print BookScouter.getCheggPrice()
wb2.save("Summer 16.xlsx")
